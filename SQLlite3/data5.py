import openpyxl

excelFile = openpyxl.load_workbook("/Users/gaegul/PycharmProjects/pythonProject/txt/score.xlsx")
aws = excelFile.active
print("++++++++++++++++++++++++++++++++++++++++++")
print("이름 국어 영어 수학 총점 평균")
print("++++++++++++++++++++++++++++++++++++++++++")
for i in aws.rows:
    index = i[0].row

    name = i[0].value

    kor = i[1].value

    eng = i[2].value

    math = i[3].value

    total = kor + eng + math
    avg = total / 3

    aws.cell(row = index,column = 5).value = total
    aws.cell(row=index, column=6).value = avg
    print('{}\t {} {}\t{}\t{}\t{:.1f}'.format(name, kor, eng, math, total, avg))
print("++++++++++++++++++++++++++++++++++++++++++")
excelFile.save('/Users/gaegul/PycharmProjects/pythonProject/txt/result.xlsx')
excelFile.close()